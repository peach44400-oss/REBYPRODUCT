# -*- coding: utf-8 -*-
"""doc/ 엑셀에서 DB에 없는 '새 날짜'만 증분 추가 (기존 데이터·수정사항 보존).

사용법: 공장에서 최신 엑셀을 받아 doc/ 에 덮어쓴 뒤
        python app/update_from_excel.py
- 완제품수불부: production.date 최대값 이후 시트만 → 생산/출고/LOT/비고
- 원재료/부재료: material_daily 최대 날짜 이후 시트만
- 원료수불부: material_usage 최대 날짜 이후 시트만
- 신규 제품/자재는 자동 등록 (기존 병합 규칙과 동일)
- import_excel.py(전체 초기화)와 달리 아무것도 지우지 않음
"""
import io
import re
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from db import connect, init_db
import import_excel as IX

# 과거/오타 표기 → 현재 제품명 (2026-07-07 중복 병합 정리 — 엑셀이 옛 표기를 계속 쓰면
# 새 제품으로 재생성되는 걸 막는다. 좌측은 엑셀에 등장하는 표기 그대로.)
PRODUCT_ALIASES = {
    # 소보로: 정식 명칭 = 소보로도넛 (40 g) — 45g 표기는 오타(2026-07-07 사용자 확정, 통합됨)
    "소보로15 골드치즈 (45 g)": "소보로도넛 골드치즈 (40 g)",
    "소보로15 바나나 (45 g)": "소보로도넛 바나나 (40 g)",
    "소보로15 블루베리 (45 g)": "소보로도넛 블루베리 (40 g)",
    "소보로15 코코넛 (45 g)": "소보로도넛 코코넛 (40 g)",
    "소보로15 코코아 (45 g)": "소보로도넛 코코아 (40 g)",
    "소보로15 골드치즈 (40 g)": "소보로도넛 골드치즈 (40 g)",
    "소보로15 바나나 (40 g)": "소보로도넛 바나나 (40 g)",
    "소보로15 블루베리 (40 g)": "소보로도넛 블루베리 (40 g)",
    "소보로15 코코넛 (40 g)": "소보로도넛 코코넛 (40 g)",
    "소보로15 코코아 (40 g)": "소보로도넛 코코아 (40 g)",
    "소보도넛 블루베리 (45 g)": "소보로도넛 블루베리 (40 g)",
    "소보로도넛 골드치즈 (45 g)": "소보로도넛 골드치즈 (40 g)",
    "소보로도넛 바나나 (45 g)": "소보로도넛 바나나 (40 g)",
    "소보로도넛 블루베리 (45 g)": "소보로도넛 블루베리 (40 g)",
    "소보로도넛 코코넛 (45 g)": "소보로도넛 코코넛 (40 g)",
    "소보로도넛 코코아 (45 g)": "소보로도넛 코코아 (40 g)",
    # 파운드: 정식 명칭 = 'X 파운드' (플랫파운드 X (80 g)·X 파운드 케이크 표기 통합, 빵아빵아는 별개)
    "플랫파운드 골드치즈 (80 g)": "치즈 파운드",
    "플랫 파운드 아몬드 (80 g)": "아몬드 파운드",
    "플랫파운드 소보로 (80 g)": "소보로 파운드",
    "치즈 파운드 케이크": "치즈 파운드",
    "아몬드 파운드 케이크": "아몬드 파운드",
    "소보로 파운드 케이크": "소보로 파운드",
    # 원료수불부 유래 변형 (병합됨)
    "맘스 우리밀": "맘스 우리밀 도넛 (45 g)",
    "맘스 우리밀 소보로": "맘스 우리밀 소보로 도넛 (45 g)",
    "맘스 우리밀 모카소보로": "맘스 우리밀 모카 소보로 도넛 (45 g)",
    "맘스 우리밀 카카오": "맘스 우리밀 카카오 도넛 (45 g)",
    "The촉촉 한오리지 널도넛": "넛파인 The촉촉한 오리지널 도넛 (40 g)",
    "The촉촉 한아몬드 도넛": "넛파인 The촉촉한 아몬드 도넛 (40 g)",
    "The촉촉 한 버터 도넛": "넛파인 The촉촉한 버터 도넛 (40 g)",
    "오븐에구 운소보로 도넛": "오븐에 구운 소보로 도넛 (이마트·트레이더스 공용)",
    "새모앤도 넛초코": "세모앤도넛 초코 (45 g)",
    "할라피뇨 페페로니 베이글 냉동생지": "할라피뇨페퍼로니_베이글_냉동생지(120g)",
    "푸르나이_우리밀소보로통밀도넛(40 g)": "푸르나이 통우리밀 소보로 도넛 (40g)",
    "푸르나이_우리밀초코통밀도넛 (40 g)": "푸르나이 통우리밀 초코 도넛 (40g)",
    "푸르나이 통밀소보로": "푸르나이 통우리밀 소보로 도넛 (40g)",
    "푸르나이 통밀 초코": "푸르나이 통우리밀 초코 도넛 (40g)",
    "푸르나이 통핑 초코": "푸르나이 통우리밀 초코 도넛 (40g)",
    "달광도넛 골즈치즈 (냉동수출)": "달광도넛 골드치즈 (냉동수출)",
}


def seed_aliases(con, cache, keyfn):
    """별칭을 캐시에 주입 — 실제 제품명(정확 매칭)이 항상 우선하도록 setdefault."""
    for alias, canon in PRODUCT_ALIASES.items():
        row = con.execute("SELECT id FROM product WHERE name=?", (canon,)).fetchone()
        if row:
            cache.setdefault(keyfn(alias), row["id"])


def update_finished(con):
    last = con.execute("SELECT COALESCE(MAX(date),'0000') FROM production").fetchone()[0]
    wb = load_workbook(IX.F_FIN, read_only=True, data_only=True)
    entries = sorted((IX.sheet_date(n), n) for n in wb.sheetnames
                     if IX.sheet_date(n) and IX.sheet_date(n) > last)
    pcache = {IX.pkey(r["name"]): r["id"] for r in con.execute("SELECT id,name FROM product")}
    seed_aliases(con, pcache, IX.pkey)
    n_prod = n_ship = n_lot = 0
    for d, sn in entries:
        grid = [list(r) for r in wb[sn].iter_rows(min_row=1, max_col=24, values_only=True)]
        cat, key = "", None
        acc, notes, lots, names = {}, {}, {}, {}
        for ri in range(5, len(grid)):
            row = grid[ri]
            b, c, packs = row[1], row[2], row[3]
            if b is not None and str(b).strip():
                cat = re.sub(r"\s+", " ", str(b).strip())
            if c is not None and str(c).strip():
                disp = re.sub(r"\s+", " ", str(c).strip().lstrip("·• "))
                if IX.TOTAL_PAT.search(disp):
                    key = None
                    continue
                key = IX.pkey(disp)
                names[key] = (disp, cat)
            if packs is not None and str(packs).strip() and key:
                a = acc.setdefault(key, [0.0, 0.0, 0.0])
                a[0] += IX.num(row[4]); a[1] += IX.num(row[5]); a[2] += IX.num(row[6])
                if row[8] is not None and str(row[8]).strip():
                    notes.setdefault(key, []).append(str(row[8]).strip())
                for kind, cols in IX.LOT_COLS.items():
                    for si, (qc, dc) in enumerate(cols):
                        exp = IX.cell_date(grid[ri + 2][qc]) if ri + 2 < len(grid) else ""
                        for off in (0, 1):
                            if ri + off >= len(grid):
                                continue
                            q = grid[ri + off][qc]
                            if isinstance(q, (int, float)) and q > 0:
                                lots.setdefault(key, []).append(
                                    (kind, si, float(q), IX.cell_date(grid[ri + off][dc]), exp))
        for k, (disp, kcat) in names.items():
            if k not in pcache:
                con.execute("INSERT OR IGNORE INTO product(name, category) VALUES(?,?)",
                            (disp, kcat))
                pcache[k] = con.execute("SELECT id FROM product WHERE name=?",
                                        (disp,)).fetchone()[0]
        con.execute("INSERT OR IGNORE INTO day_record(date) VALUES(?)", (d,))
        for k, a in acc.items():
            if k not in pcache:
                continue
            note = "; ".join(notes.get(k, []))
            if a[2] > 0 or note:
                con.execute("""INSERT OR REPLACE INTO production(date,product_id,prod_qty,note)
                    VALUES(?,?,?,?)""", (d, pcache[k], a[2], note))
                n_prod += 1
            if a[1] > 0:
                con.execute("INSERT INTO shipment(date,product_id,qty) VALUES(?,?,?)",
                            (d, pcache[k], a[1]))
                n_ship += 1
        for k, ls in lots.items():
            if k not in pcache:
                continue
            for kind, si, q, made, exp in ls:
                con.execute("""INSERT INTO lot_snapshot(date,product_id,kind,slot,qty,made_date,expiry)
                    VALUES(?,?,?,?,?,?,?)""", (d, pcache[k], kind, si, q, made, exp))
                n_lot += 1
    wb.close()
    print(f"[완제품] 기준 {last} 이후 새 시트 {len(entries)}개 → 생산 {n_prod} · 출고 {n_ship} · LOT {n_lot}행 추가")


def update_materials(con, path, kind):
    last = con.execute("""SELECT COALESCE(MAX(md.date),'0000') FROM material_daily md
        JOIN material m ON m.id=md.material_id WHERE m.kind=?""", (kind,)).fetchone()[0]
    wb = load_workbook(path, read_only=True, data_only=True)
    entries = sorted((IX.sheet_date(n), n) for n in wb.sheetnames
                     if IX.sheet_date(n) and IX.sheet_date(n) > last)
    mcache = {IX.mkey(r["name"]): r["id"] for r in
              con.execute("SELECT id,name FROM material WHERE kind=?", (kind,))}
    n_daily = 0
    for d, sn in entries:
        grid = [list(r) for r in wb[sn].iter_rows(min_row=1, max_col=16, values_only=True)]
        if kind == "raw":
            hri, C = IX.detect_header(grid)
            if hri is None or "name" not in C or "real" not in C or "prev" not in C:
                continue
            rows_iter = []
            for row in grid[hri + 1:]:
                name = row[C["name"]] if C["name"] < len(row) else None
                if name is None or not str(name).strip() or IX.is_num_like(name):
                    continue
                name = re.sub(r"\s+", " ", str(name).strip())
                if IX.TOTAL_PAT.search(name) or IX.nrm(name) in ("품명", "(kg)"):
                    continue
                prev = IX.num(row[C["prev"]])
                inq = IX.num(row[C["inq"]]) if "inq" in C else 0.0
                real = IX.num(row[C["real"]])
                uv = row[C["used"]] if "used" in C and C["used"] < len(row) else None
                used = IX.num(uv) if isinstance(uv, (int, float)) else prev + inq - real
                order = row[C["order"]] if "order" in C and C["order"] < len(row) else ""
                rows_iter.append((name, prev, inq, real, used,
                                  str(order).strip() if order is not None else "", 0.0))
        else:
            rows_iter = []
            for row in grid[6:]:
                name = row[1] if len(row) > 1 else None
                if name is None or not str(name).strip() or IX.is_num_like(name):
                    continue
                name = re.sub(r"\s+", " ", str(name).strip())
                if IX.TOTAL_PAT.search(name):
                    continue
                prev, inq, real = IX.num(row[7]), IX.num(row[5]), IX.num(row[2])
                uv = row[6]
                used = IX.num(uv) if isinstance(uv, (int, float)) else prev + inq - real
                order = str(row[3]).strip() if row[3] is not None else ""
                rows_iter.append((name, prev, inq, real, used, order, IX.num(row[4])))
        if not rows_iter:
            continue
        con.execute("INSERT OR IGNORE INTO day_record(date) VALUES(?)", (d,))
        for name, prev, inq, real, used, order, oqty in rows_iter:
            k = IX.mkey(name)
            if k not in mcache:
                con.execute("INSERT OR IGNORE INTO material(kind,name,unit) VALUES(?,?,?)",
                            (kind, name, "kg" if kind == "raw" else "ea"))
                mcache[k] = con.execute("SELECT id FROM material WHERE kind=? AND name=?",
                                        (kind, name)).fetchone()[0]
            con.execute("""INSERT OR REPLACE INTO material_daily
                (date,material_id,prev_qty,in_qty,real_qty,used_qty,order_date,order_qty)
                VALUES(?,?,?,?,?,?,?,?)""",
                        (d, mcache[k], prev, inq, real, used, order, oqty))
            n_daily += 1
    wb.close()
    label = "원재료" if kind == "raw" else "부재료"
    print(f"[{label}] 기준 {last} 이후 새 시트 {len(entries)}개 → 일일 {n_daily}행 추가")


def update_usage(con):
    last = con.execute("SELECT COALESCE(MAX(date),'0000') FROM material_usage").fetchone()[0]
    wb = load_workbook(IX.F_USE, read_only=True, data_only=True)
    entries = sorted((IX.sheet_date(n), n) for n in wb.sheetnames
                     if IX.sheet_date(n) and IX.sheet_date(n) > last)
    prod_norm = {IX.nrm_prod(r["name"]): r["id"] for r in con.execute("SELECT id,name FROM product")}
    seed_aliases(con, prod_norm, IX.nrm_prod)
    real_norms = [(IX.nrm_prod(r["name"]), r["id"]) for r in
                  con.execute("SELECT id,name FROM product WHERE note=''")]
    mat_norm = {IX.mkey(r["name"]): r["id"] for r in
                con.execute("SELECT id,name FROM material WHERE kind='raw'")}
    n_cells = 0
    for d, sn in entries:
        ws = wb[sn]
        grid = [list(r) for r in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                              max_col=ws.max_column, values_only=True)]
        if len(grid) < 10:
            continue
        hdr = grid[7]
        end = None
        for i, h in enumerate(hdr):
            if h and ("사용원료" in IX.nrm(h) or "금일사용" in IX.nrm(h)):
                end = i
                break
        if end is None:
            end = len(hdr)
        pcols = []
        for i in range(4, end):
            h = hdr[i]
            if not h or not IX.nrm(h):
                continue
            k = IX.nrm_prod(h)
            pid = prod_norm.get(k)
            if pid is None and k:
                cands = [rid for rn, rid in real_norms if k in rn]
                if len(cands) == 1:
                    pid = cands[0]
                    prod_norm[k] = pid
            if pid is None:
                disp = re.sub(r"\s+", " ", str(h).replace("\n", " ").strip())
                con.execute("INSERT OR IGNORE INTO product(name, note) VALUES(?, '원료수불부 유래')",
                            (disp,))
                pid = con.execute("SELECT id FROM product WHERE name=?", (disp,)).fetchone()[0]
                prod_norm[k] = pid
            pcols.append((i, pid))
        con.execute("INSERT OR IGNORE INTO day_record(date) VALUES(?)", (d,))
        for row in grid[9:]:
            name = row[1]
            if name is None or not str(name).strip() or IX.is_num_like(name):
                continue
            name = re.sub(r"\s+", " ", str(name).strip())
            if IX.TOTAL_PAT.search(name):
                continue
            k = IX.mkey(name)
            mid = mat_norm.get(k)
            if mid is None:
                con.execute("INSERT OR IGNORE INTO material(kind,name,unit,note)"
                            " VALUES('raw',?,?,'원료수불부 유래')", (name, "kg"))
                mid = con.execute("SELECT id FROM material WHERE kind='raw' AND name=?",
                                  (name,)).fetchone()[0]
                mat_norm[k] = mid
            for ci, pid in pcols:
                v = row[ci] if ci < len(row) else None
                if isinstance(v, (int, float)) and v > 0:
                    con.execute("""INSERT OR REPLACE INTO material_usage
                        (date,material_id,product_id,qty) VALUES(?,?,?,?)""",
                                (d, mid, pid, float(v)))
                    n_cells += 1
    wb.close()
    print(f"[원료수불부] 기준 {last} 이후 새 시트 {len(entries)}개 → 사용량 {n_cells}셀 추가")


def main():
    init_db()
    con = connect()
    update_finished(con)
    update_materials(con, IX.F_RAW, "raw")
    update_materials(con, IX.F_SUB, "sub")
    update_usage(con)
    con.execute("INSERT INTO audit_log(action,detail) VALUES('excel_update','증분 업데이트 실행')")
    con.commit()
    r = con.execute("SELECT MAX(date) FROM production").fetchone()[0]
    r2 = con.execute("SELECT MAX(date) FROM material_daily").fetchone()[0]
    print(f"\n완료 — 생산 최신일 {r} · 자재 최신일 {r2}")
    con.close()


if __name__ == "__main__":
    main()
