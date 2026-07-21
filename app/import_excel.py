# -*- coding: utf-8 -*-
"""doc/ 폴더의 엑셀 4종을 martin_stock.db로 임포트 (재실행 시 전체 갱신 — 사용 시작 후 금지).

원본 → DB 매핑:
- ★ 전제품완제품수불부: product + production(생산·비고)/shipment(출고) + lot_snapshot(우측 LOT 블록:
  생산일자별 재고/출고 + 소비기한) + 기초재고
- 5.원재료: material(raw) + material_daily(현행양식 2026~) + material_usage_type(용도별 배합량,
  구양식 2025-09~12 포함 — 헤더 자동탐지)
- 5.부재료: material(sub) + material_daily(+발주량) + 생산가능 수식에서 mult/per 추출
- ★전제품_원료수불부: material_usage(자재×제품 매트릭스) + 유래 자재의 material_daily
- 생산관리보고서: 양식(예시 데이터) → 라인/거래처 시드만

미활용 데이터와 사유는 README 참고.
"""
import io
import re
import sys
import datetime as dt
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from db import connect, init_db, DB_PATH

DOC = Path(__file__).resolve().parent.parent / "doc"
F_FIN = DOC / "★ 전제품완제품수불부_선행요건.xlsx"
F_RAW = DOC / "5.원재료 사용량 및 재고.xlsx"
F_SUB = DOC / "5.부재료 사용량 및 재고.xlsx"
F_USE = DOC / "★전제품_원료수불부_선행요건.xlsx"

TOTAL_PAT = re.compile(r"(TOTAL|합계|소계|총계)", re.I)
DATE_PAT = re.compile(r"(\d{4})[-./](\d{1,2})[-./](\d{1,2})")
MATDAILY_FROM = "2026-01-01"   # material_daily 는 현행양식(실재고 기반)만


def nrm(s):
    return re.sub(r"\s+", "", str(s or ""))


def pkey(s):
    """제품 병합 키: 공백·_·· 제거 (무게/맛 차이는 분리 유지)."""
    return re.sub(r"[\s_·•]", "", str(s or ""))


def nrm_prod(s):
    return re.sub(r"\([^)]*g\)", "", pkey(s), flags=re.I)


def mkey(s):
    """자재 병합 키: 공백·괄호·_·· 제거."""
    return re.sub(r"[\s()（）_·•]", "", str(s or ""))


def sheet_date(name):
    t = str(name).strip()
    for fmt in ("%Y-%m-%d", "%y-%m-%d", "%y.%m.%d"):
        try:
            return dt.datetime.strptime(t, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def cell_date(v):
    """셀 값 → ISO 날짜 (datetime / '2026-05-29' / '소비 : 2026-08-28' 등)."""
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    m = DATE_PAT.search(str(v))
    if m:
        y, mo, d = m.groups()
        try:
            return dt.date(int(y), int(mo), int(d)).isoformat()
        except ValueError:
            return ""
    return ""


def num(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return 0.0


def is_num_like(v):
    try:
        float(str(v).replace(",", ""))
        return True
    except (ValueError, TypeError):
        return False


def wipe(con):
    for t in ("material_usage_type", "lot_snapshot", "material_usage", "material_daily",
              "staffing_member", "staffing", "shipment", "production", "day_record",
              "opening_stock", "bom", "material", "product", "partner", "staff", "line",
              "audit_log"):
        con.execute(f"DELETE FROM {t}")


def seed(con):
    for n, p in [("1LINE", "배합/성형"), ("2LINE", "성형/포장"), ("3LINE", "포장/검수"),
                 ("품질관리(QC)", "검사")]:
        con.execute("INSERT INTO line(name, process) VALUES(?,?)", (n, p))
    for n, t in [("이마트", "판매처"), ("트레이더스", "판매처"), ("급식업체", "판매처"),
                 ("홈쇼핑", "판매처"), ("온라인 마켓", "판매처")]:
        con.execute("INSERT INTO partner(name, type) VALUES(?,?)", (n, t))


# ── 1) 완제품 수불부 (+LOT 블록, 비고) ─────────────
LOT_COLS = {"out": [(12, 13), (14, 15), (16, 17)],   # (수량, 생산일자) 0-based: M/N O/P Q/R
            "stock": [(18, 19), (20, 21), (22, 23)]}  # S/T U/V W/X


def import_finished(con):
    wb = load_workbook(F_FIN, read_only=True, data_only=True)
    entries = sorted((sheet_date(n), n) for n in wb.sheetnames if sheet_date(n))
    products = {}
    first = True
    n_prod = n_ship = n_lot = 0

    for d, sn in entries:
        grid = [list(r) for r in wb[sn].iter_rows(min_row=1, max_col=24, values_only=True)]
        cat = ""
        key = None
        acc, notes, lots = {}, {}, {}
        for ri in range(5, len(grid)):
            row = grid[ri]
            b, c, packs = row[1], row[2], row[3]
            if b is not None and str(b).strip():
                cat = re.sub(r"\s+", " ", str(b).strip())
            if c is not None and str(c).strip():
                disp = re.sub(r"\s+", " ", str(c).strip().lstrip("·• "))
                if TOTAL_PAT.search(disp):
                    key = None
                    continue
                key = pkey(disp)
                if key not in products:
                    products[key] = {"name": disp, "cat": cat, "packs": set(),
                                     "sort": len(products)}
                elif cat and not products[key]["cat"]:
                    products[key]["cat"] = cat
            if packs is not None and str(packs).strip() and key:
                products[key]["packs"].add(str(packs).strip())
                a = acc.setdefault(key, [0.0, 0.0, 0.0])
                a[0] += num(row[4]); a[1] += num(row[5]); a[2] += num(row[6])
                if row[8] is not None and str(row[8]).strip():
                    notes.setdefault(key, []).append(str(row[8]).strip())
                # LOT 블록: 슬롯값 2줄(ri, ri+1) + 소비기한(ri+2, 수량열)
                for kind, cols in LOT_COLS.items():
                    for si, (qc, dc) in enumerate(cols):
                        exp = cell_date(grid[ri + 2][qc]) if ri + 2 < len(grid) else ""
                        for off in (0, 1):
                            if ri + off >= len(grid):
                                continue
                            q = grid[ri + off][qc]
                            if isinstance(q, (int, float)) and q > 0:
                                made = cell_date(grid[ri + off][dc])
                                lots.setdefault(key, []).append((kind, si, float(q), made, exp))

        for k, meta in products.items():
            con.execute(
                "INSERT OR IGNORE INTO product(name, category, pack_sizes, sort) VALUES(?,?,?,?)",
                (meta["name"], meta["cat"], ",".join(sorted(meta["packs"])), meta["sort"]))
            con.execute("UPDATE product SET category=? WHERE name=? AND category=''",
                        (meta["cat"], meta["name"]))
        pid = {pkey(r["name"]): r["id"] for r in con.execute("SELECT id,name FROM product")}

        if first:
            for k, a in acc.items():
                con.execute("INSERT OR REPLACE INTO opening_stock VALUES('product',?,?,?)",
                            (pid[k], d, a[0]))
            first = False

        con.execute("INSERT OR REPLACE INTO day_record(date) VALUES(?)", (d,))
        for k, a in acc.items():
            note = "; ".join(notes.get(k, []))
            if a[2] > 0 or note:
                con.execute("""INSERT OR REPLACE INTO production
                    (date,product_id,prod_qty,note) VALUES(?,?,?,?)""",
                            (d, pid[k], a[2], note))
                n_prod += 1
            if a[1] > 0:
                con.execute("INSERT INTO shipment(date,product_id,qty) VALUES(?,?,?)",
                            (d, pid[k], a[1]))
                n_ship += 1
        for k, ls in lots.items():
            for kind, si, q, made, exp in ls:
                con.execute("""INSERT INTO lot_snapshot(date,product_id,kind,slot,qty,made_date,expiry)
                    VALUES(?,?,?,?,?,?,?)""", (d, pid[k], kind, si, q, made, exp))
                n_lot += 1
    wb.close()
    ds = [d for d, _ in entries]
    print(f"[완제품] 제품 {len(products)}종, {ds[0]}~{ds[-1]}, 생산 {n_prod} · 출고 {n_ship} · LOT {n_lot}행")


# ── 2) 원재료 (헤더 자동탐지: 현행+구양식) ─────────
RAW_HEADERS = {
    "name": ["품명"], "unit": ["단위"], "real": ["실재고"], "order": ["발주"],
    "inq": ["입고량", "입고"], "used": ["금일사용량"], "prev": ["전일재고", "전일"],
    "t_도넛": ["도넛"], "t_소보로토핑": ["소보로토핑"], "t_단백질": ["단백질"],
    "t_추가": ["추가사용"], "t_테스트": ["테스트"],
}


def detect_header(grid):
    """'품명' 이 있는 행을 헤더로 보고 컬럼 인덱스 매핑."""
    for ri in range(2, min(8, len(grid))):
        row = grid[ri]
        if any(nrm(c) == "품명" for c in row if c):
            cols = {}
            for ci, c in enumerate(row):
                if not c:
                    continue
                t = nrm(c)
                for k, pats in RAW_HEADERS.items():
                    if k not in cols and any(p in t for p in pats):
                        cols[k] = ci
            return ri, cols
    return None, {}


def import_raw(con):
    wb = load_workbook(F_RAW, read_only=True, data_only=True)
    entries = sorted((sheet_date(n), n) for n in wb.sheetnames if sheet_date(n))
    skipped_noyear = len(wb.sheetnames) - len(entries)
    mats = {}
    first_daily = True
    n_daily = n_type = 0
    daily_dates, type_dates = [], []

    for d, sn in entries:
        grid = [list(r) for r in wb[sn].iter_rows(min_row=1, max_col=16, values_only=True)]
        hri, C = detect_header(grid)
        if hri is None or "name" not in C:
            continue
        has_daily = d >= MATDAILY_FROM and "real" in C and "prev" in C
        rows_out = []
        for row in grid[hri + 1:]:
            name = row[C["name"]] if C["name"] < len(row) else None
            if name is None or not str(name).strip() or is_num_like(name):
                continue
            name = re.sub(r"\s+", " ", str(name).strip())
            if TOTAL_PAT.search(name) or nrm(name) in ("품명", "(kg)"):
                continue
            rows_out.append((name, row))
        if not rows_out:
            continue

        for name, row in rows_out:
            k = mkey(name)
            if k not in mats:
                unitspec = row[C["unit"]] if "unit" in C and C["unit"] < len(row) else None
                spec = f"{unitspec}kg" if unitspec not in (None, "") else ""
                con.execute(
                    "INSERT OR IGNORE INTO material(kind,name,spec,unit,sort) VALUES(?,?,?,?,?)",
                    ("raw", name, spec, "kg", len(mats)))
                mats[k] = con.execute(
                    "SELECT id FROM material WHERE kind='raw' AND name=?", (name,)).fetchone()[0]

        con.execute("INSERT OR REPLACE INTO day_record(date) VALUES(?)", (d,))
        if has_daily:
            daily_dates.append(d)
            for name, row in rows_out:
                prev = num(row[C["prev"]]) if C["prev"] < len(row) else 0.0
                inq = num(row[C["inq"]]) if "inq" in C and C["inq"] < len(row) else 0.0
                real = num(row[C["real"]]) if C["real"] < len(row) else 0.0
                used_v = row[C["used"]] if "used" in C and C["used"] < len(row) else None
                used = num(used_v) if isinstance(used_v, (int, float)) else prev + inq - real
                order = row[C["order"]] if "order" in C and C["order"] < len(row) else ""
                if first_daily:
                    con.execute("INSERT OR REPLACE INTO opening_stock VALUES('material',?,?,?)",
                                (mats[mkey(name)], d, prev))
                con.execute("""INSERT OR REPLACE INTO material_daily
                    (date,material_id,prev_qty,in_qty,real_qty,used_qty,order_date)
                    VALUES(?,?,?,?,?,?,?)""",
                            (d, mats[mkey(name)], prev, inq, real, used,
                             str(order).strip() if order is not None else ""))
                n_daily += 1
            first_daily = False
        # 용도별 배합량 (신·구양식 공통)
        tcols = [(k[2:], C[k]) for k in C if k.startswith("t_")]
        if tcols:
            wrote = False
            for name, row in rows_out:
                for tname, ci in tcols:
                    v = row[ci] if ci < len(row) else None
                    if isinstance(v, (int, float)) and v > 0:
                        con.execute("""INSERT OR REPLACE INTO material_usage_type
                            (date,material_id,type,qty) VALUES(?,?,?,?)""",
                                    (d, mats[mkey(name)], tname, float(v)))
                        n_type += 1
                        wrote = True
            if wrote:
                type_dates.append(d)
    wb.close()
    print(f"[원재료] 자재 {len(mats)}종 · 일일 {n_daily}행({daily_dates[0]}~{daily_dates[-1]})"
          f" · 용도별 {n_type}행({type_dates[0]}~{type_dates[-1]})"
          f" · 연도불명 시트 {skipped_noyear}개 제외")


# ── 3) 부재료 (+발주량, 생산가능 수식 추출) ────────
def import_sub(con):
    wb = load_workbook(F_SUB, read_only=True, data_only=True)
    entries = sorted((sheet_date(n), n) for n in wb.sheetnames
                     if sheet_date(n) and sheet_date(n) >= MATDAILY_FROM)
    mats = {}
    first = True
    n_daily = 0
    for d, sn in entries:
        rows_out = []
        for row in wb[sn].iter_rows(min_row=7, max_col=10, values_only=True):
            name = row[1]
            if name is None or not str(name).strip() or is_num_like(name):
                continue
            name = re.sub(r"\s+", " ", str(name).strip())
            if TOTAL_PAT.search(name):
                continue
            rows_out.append((name, row))
        for name, row in rows_out:
            k = mkey(name)
            if k not in mats:
                con.execute(
                    "INSERT OR IGNORE INTO material(kind,name,unit,sort) VALUES(?,?,?,?)",
                    ("sub", name, "ea", len(mats)))
                mats[k] = con.execute(
                    "SELECT id FROM material WHERE kind='sub' AND name=?", (name,)).fetchone()[0]
        con.execute("INSERT OR REPLACE INTO day_record(date) VALUES(?)", (d,))
        for name, row in rows_out:
            real, order, oqty, inq, used_v, prev = row[2], row[3], row[4], row[5], row[6], row[7]
            prev, inq, real = num(prev), num(inq), num(real)
            used = num(used_v) if isinstance(used_v, (int, float)) else prev + inq - real
            if first:
                con.execute("INSERT OR REPLACE INTO opening_stock VALUES('material',?,?,?)",
                            (mats[mkey(name)], d, prev))
            con.execute("""INSERT OR REPLACE INTO material_daily
                (date,material_id,prev_qty,in_qty,real_qty,used_qty,order_date,order_qty)
                VALUES(?,?,?,?,?,?,?,?)""",
                        (d, mats[mkey(name)], prev, inq, real, used,
                         str(order).strip() if order is not None else "", num(oqty)))
            n_daily += 1
        first = False
    wb.close()

    # 생산가능 수식(I=C*mult, J=I/per)에서 환산계수 추출 → material.prod_mult/per
    n_formula = 0
    wb2 = load_workbook(F_SUB, read_only=True, data_only=False)
    last = sorted((sheet_date(n), n) for n in wb2.sheetnames if sheet_date(n))[-1][1]
    for row in wb2[last].iter_rows(min_row=7, max_col=10, values_only=True):
        name, fi, fj = row[1], row[8], row[9]
        if name is None or not str(name).strip():
            continue
        k = mkey(re.sub(r"\s+", " ", str(name).strip()))
        if k not in mats:
            continue
        mult = per = None
        if isinstance(fi, str):
            m = re.search(r"\*\s*([\d.]+)", fi)
            if m:
                mult = float(m.group(1))
        if isinstance(fj, str):
            m = re.search(r"/\s*([\d.]+)", fj)
            if m:
                per = float(m.group(1))
        if mult or per:
            con.execute("UPDATE material SET prod_mult=?, prod_per=? WHERE id=?",
                        (mult, per, mats[k]))
            n_formula += 1
    wb2.close()
    ds = [d for d, _ in entries]
    print(f"[부재료] 자재 {len(mats)}종 · 일일 {n_daily}행({ds[0]}~{ds[-1]})"
          f" · 생산가능 환산계수 {n_formula}종 추출")


# ── 4) 원료수불부 (사용량 매트릭스 + 유래자재 일일) ──
def import_usage(con):
    wb = load_workbook(F_USE, read_only=True, data_only=True)
    prod_norm = {nrm_prod(r["name"]): r["id"] for r in con.execute("SELECT id,name FROM product")}
    # 포함 매칭용: 실제 제품(수불부 유래 아님) — 유일하게 포함될 때만 매칭
    real_norms = [(nrm_prod(r["name"]), r["id"]) for r in
                  con.execute("SELECT id,name FROM product WHERE note=''")]
    raw_ids = {mkey(r["name"]): r["id"] for r in
               con.execute("SELECT id,name FROM material WHERE kind='raw'")}
    book_only = {}   # 이 장부에만 있는 자재 → material_daily 도 여기서
    new_products, new_materials, matched_p = set(), set(), set()
    n_cells = n_daily = 0
    dates = []

    for sn in wb.sheetnames:
        d = sheet_date(sn)
        if not d:
            continue
        dates.append(d)
        ws = wb[sn]
        grid = [list(r) for r in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                              max_col=ws.max_column, values_only=True)]
        if len(grid) < 10:
            continue
        hdr = grid[7]
        end = use_ci = cur_ci = None
        for i, h in enumerate(hdr):
            t = nrm(h) if h else ""
            if end is None and h and ("사용원료" in t or "금일사용" in t):
                end = i
            if h and "금일사용" in t:
                use_ci = i
            if h and "현재고" in t.replace(" ", ""):
                cur_ci = i
        if end is None:
            end = len(hdr)
        pcols = []
        for i in range(4, end):
            h = hdr[i]
            if not h or not nrm(h):
                continue
            k = nrm_prod(h)
            pid = prod_norm.get(k)
            if pid is None and k:
                # 정확 매칭 실패 → 실제 제품명에 유일하게 포함되면 그 제품으로
                cands = [rid for rn, rid in real_norms if k in rn]
                if len(cands) == 1:
                    pid = cands[0]
                    prod_norm[k] = pid
                    matched_p.add(k)
            if pid is None:
                disp = re.sub(r"\s+", " ", str(h).replace("\n", " ").strip())
                con.execute(
                    "INSERT OR IGNORE INTO product(name, note, sort) VALUES(?,?,?)",
                    (disp, "원료수불부 유래", 900 + len(new_products)))
                pid = con.execute("SELECT id FROM product WHERE name=?", (disp,)).fetchone()[0]
                prod_norm[k] = pid
                new_products.add(disp)
            else:
                matched_p.add(k)
            pcols.append((i, pid))

        con.execute("INSERT OR REPLACE INTO day_record(date) VALUES(?)", (d,))
        for row in grid[9:]:
            name = row[1]
            if name is None or not str(name).strip() or is_num_like(name):
                continue
            name = re.sub(r"\s+", " ", str(name).strip())
            if TOTAL_PAT.search(name):
                continue
            k = mkey(name)
            mid = raw_ids.get(k)
            if mid is None:
                con.execute(
                    "INSERT OR IGNORE INTO material(kind,name,unit,note,sort) VALUES(?,?,?,?,?)",
                    ("raw", name, "kg", "원료수불부 유래", 900 + len(new_materials)))
                mid = con.execute(
                    "SELECT id FROM material WHERE kind='raw' AND name=?", (name,)).fetchone()[0]
                raw_ids[k] = mid
                new_materials.add(name)
                book_only[k] = mid
            for ci, pid in pcols:
                v = row[ci] if ci < len(row) else None
                if isinstance(v, (int, float)) and v > 0:
                    con.execute(
                        "INSERT OR REPLACE INTO material_usage(date,material_id,product_id,qty)"
                        " VALUES(?,?,?,?)", (d, mid, pid, float(v)))
                    n_cells += 1
            # 이 장부에만 있는 자재: 전일(2)/입고(3)/사용/현재고 → material_daily
            if k in book_only and cur_ci:
                prev, inq = num(row[2]), num(row[3])
                used = num(row[use_ci]) if use_ci and use_ci < len(row) else 0.0
                real = num(row[cur_ci]) if cur_ci < len(row) else prev + inq - used
                if prev or inq or used or real:
                    con.execute("""INSERT OR REPLACE INTO material_daily
                        (date,material_id,prev_qty,in_qty,real_qty,used_qty)
                        VALUES(?,?,?,?,?,?)""", (d, mid, prev, inq, real, used))
                    n_daily += 1
    wb.close()
    dates.sort()
    print(f"[원료수불부] {dates[0]}~{dates[-1]} · 사용량셀 {n_cells} · 유래자재 일일 {n_daily}행")
    print(f"  제품 매칭 {len(matched_p)} / 신규 {len(new_products)} · 자재 신규 {len(new_materials)}")


def report(con):
    print("\n=== DB 요약 ===", DB_PATH)
    for t in ("product", "material", "partner", "line", "day_record", "production",
              "shipment", "material_daily", "material_usage", "material_usage_type",
              "lot_snapshot"):
        n = con.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
        print(f"  {t:20s} {n:>8,}")
    r = con.execute("SELECT MIN(date), MAX(date) FROM day_record").fetchone()
    print(f"  기간: {r[0]} ~ {r[1]}")


def tag_legacy(con):
    """구양식(2025)에서만 등장한 원재료 = 현행 재고·사용처 기록이 전혀 없음 → 중단 처리."""
    con.execute("""UPDATE material SET status='중단', note='구양식(2025) 명칭'
        WHERE kind='raw' AND note=''
          AND id NOT IN (SELECT DISTINCT material_id FROM material_daily)
          AND id NOT IN (SELECT DISTINCT material_id FROM material_usage)""")
    n = con.execute("SELECT COUNT(*) FROM material WHERE note='구양식(2025) 명칭'").fetchone()[0]
    print(f"[정리] 구양식 명칭 자재 {n}종 → 중단 표시")


def main():
    init_db()
    con = connect()
    wipe(con)
    seed(con)
    import_finished(con)
    import_raw(con)
    import_sub(con)
    import_usage(con)
    tag_legacy(con)
    con.commit()
    report(con)
    con.close()


if __name__ == "__main__":
    main()
