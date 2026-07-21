# -*- coding: utf-8 -*-
"""bom 테이블 → 배합비 통합정리 엑셀 (doc/배합비/배합비_통합정리.xlsx).

사용법: python app/export_bom.py
제품×블록(반죽/토핑)×자재 전체를 한 시트에 정리 — 1배합당 소요량은 배합비 엑셀 원본 수치.
"""
import io
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, str(Path(__file__).resolve().parent))
from db import connect

OUT = Path(__file__).resolve().parent.parent / "doc" / "배합비" / "배합비_통합정리.xlsx"

NAVY = "1F3864"
LIGHT = "D9E2F3"
THIN = Side(style="thin", color="BBBBBB")
TOP = Side(style="medium", color="1F3864")


def main():
    con = connect()
    rows = con.execute("""
        SELECT p.name pname, p.batch_yield,
               CASE WHEN b.block='' THEN '(실측/수동)' ELSE b.block END block,
               m.name mname, m.kind,
               b.batch_qty, b.block_yield, b.qty_per_unit, b.note
        FROM bom b
        JOIN product p ON p.id=b.product_id
        JOIN material m ON m.id=b.material_id
        ORDER BY p.name, CASE b.block WHEN '반죽' THEN 0 WHEN '토핑' THEN 1 ELSE 2 END, b.id
    """).fetchall()
    con.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "배합비 통합정리"

    ws["A1"] = "배합비 통합 정리표"
    ws["A1"].font = Font(size=15, bold=True, color=NAVY)
    ws["A2"] = "1배합당 소요량 = 배합비 엑셀 원본 수치 · 1개당 = 1배합당 ÷ 그 배합 생산수량 · 반죽/토핑은 수율이 다른 별도 배합"
    ws["A2"].font = Font(size=9, color="666666")

    headers = ["제품명", "배합", "자재명", "구분", "1배합당 소요량 (g)",
               "1배합 생산수량 (개)", "1개당 (g)", "1,000개당 (kg)", "비고"]
    HR = 4
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=HR, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

    r = HR + 1
    prev_prod = None
    for row in rows:
        first = row["pname"] != prev_prod
        vals = [row["pname"] if first else "", row["block"], row["mname"],
                "원재료" if row["kind"] == "raw" else "부재료",
                row["batch_qty"] or None,
                row["block_yield"] or (row["batch_yield"] or None),
                round(row["qty_per_unit"], 4),
                round(row["qty_per_unit"], 4),  # g/개 × 1,000개 ÷ 1,000(g→kg) = 수치 동일
                row["note"] or ""]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.border = Border(top=TOP if first and ci <= len(headers) else THIN,
                              bottom=THIN, left=THIN, right=THIN)
            if ci == 1:
                c.font = Font(bold=True, size=10)
            elif ci == 2:
                c.alignment = Alignment(horizontal="center")
                if v == "토핑":
                    c.fill = PatternFill("solid", fgColor=LIGHT)
            if ci in (5, 6):
                c.number_format = "#,##0"
            if ci in (7, 8):
                c.number_format = "#,##0.0000"
        prev_prod = row["pname"]
        r += 1

    widths = [34, 10, 20, 8, 17, 17, 11, 13, 24]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + ci)].width = w
    ws.freeze_panes = f"A{HR + 1}"
    ws.auto_filter.ref = f"A{HR}:I{r - 1}"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"저장: {OUT}  ({r - HR - 1}행, 제품 {len({x['pname'] for x in rows})}종)")


if __name__ == "__main__":
    main()
