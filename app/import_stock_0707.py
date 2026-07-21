# -*- coding: utf-8 -*-
"""7/7 실사 기준 재고 재설정 (2026-07-09 사용자 요청).

원천: doc/5.원재료 사용량 및 재고.xlsx · doc/5.부재료 사용량 및 재고.xlsx — 각 '26-07-07' 시트.
사용자 지시: '금일 실재고'만 참고 (전일재고·사용량 등은 무시).

수행:
1) DB 백업
2) 2026-07-07 material_daily(원료수불부 유래 72건)·material_usage(17셀) 삭제
3) 시트의 금일 실재고 → material_daily 2026-07-07 실사(prev=real, used=0, src=manual)
   미매칭 자재는 신규 생성 (kind는 시트 기준)
4) 2026-07-08 자동차감(auto) 행의 전일재고를 새 7/7 실재고로 재계산
5) 부재료 생산가능 환산계수(prod_mult=수량/재고, prod_per=수량/횟수) 갱신
"""
import sys, io, re, shutil, sqlite3, datetime as dt
from pathlib import Path

if hasattr(sys.stdout, "buffer") and (sys.stdout.encoding or "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import openpyxl

BASE = Path(__file__).resolve().parent.parent
DB = BASE / "martin_stock.db"
DATE = "2026-07-07"
SHEET = "26-07-07"


def mkey(s):
    return re.sub(r"[\s()（）_·•]", "", str(s or ""))


SKIP = re.compile(r"^(sub\s*total|total|소계|합계)$", re.I)


def num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def parse_raw():
    """원재료 시트: C=품명, E=금일 실재고."""
    wb = openpyxl.load_workbook(BASE / "doc" / "5.원재료 사용량 및 재고.xlsx",
                                read_only=True, data_only=True)
    ws = wb[SHEET]
    out = []
    for r in range(6, ws.max_row + 1):
        name = ws.cell(r, 3).value
        if not name or not str(name).strip() or SKIP.match(str(name).strip()):
            continue
        real = num(ws.cell(r, 5).value)
        if real is None:
            continue
        out.append((str(name).strip(), real))
    wb.close()
    return out


def parse_sub():
    """부재료 시트: B=품명, C=금일 실재고, I=생산가능수량, J=생산가능횟수."""
    wb = openpyxl.load_workbook(BASE / "doc" / "5.부재료 사용량 및 재고.xlsx",
                                read_only=True, data_only=True)
    ws = wb[SHEET]
    out = []
    for r in range(7, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name or not str(name).strip() or SKIP.match(str(name).strip()):
            continue
        real = num(ws.cell(r, 3).value)
        if real is None:
            continue
        out.append((str(name).strip(), real,
                    num(ws.cell(r, 9).value), num(ws.cell(r, 10).value)))
    wb.close()
    return out


def main():
    bak = BASE / "백업" / f"백업_{dt.date.today():%Y%m%d}_0707실사재설정전.db"
    shutil.copy2(DB, bak)
    print("백업:", bak.name)

    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row

    # 자재 매칭 맵: mkey → {kind: id} (같은 이름이 raw/sub 양쪽에 있으면 시트 kind 우선)
    mats = {}
    for r in con.execute("SELECT id, kind, name FROM material"):
        mats.setdefault(mkey(r["name"]), {})[r["kind"]] = r["id"]

    def find_or_create(name, kind):
        d = mats.get(mkey(name))
        if d:
            mid = d.get(kind) or next(iter(d.values()))   # 같은 kind 우선, 없으면 반대 kind
            return mid, False
        cur = con.execute(
            "INSERT INTO material(kind, name, unit, note) VALUES(?,?,?,?)",
            (kind, name, "kg" if kind == "raw" else "ea", "7/7 실사 유래"))
        mats.setdefault(mkey(name), {})[kind] = cur.lastrowid
        return cur.lastrowid, True

    n_del_md = con.execute("DELETE FROM material_daily WHERE date=?", (DATE,)).rowcount
    n_del_mu = con.execute("DELETE FROM material_usage WHERE date=?", (DATE,)).rowcount
    print(f"삭제: {DATE} 자재기록 {n_del_md}건 · 제품별사용 {n_del_mu}건")

    created, n_raw, n_sub, n_mult = [], 0, 0, 0
    for name, real in parse_raw():
        mid, new = find_or_create(name, "raw")
        if new:
            created.append(("raw", name))
        con.execute("""INSERT OR REPLACE INTO material_daily
            (date, material_id, prev_qty, in_qty, real_qty, used_qty, src)
            VALUES(?,?,?,?,?,?,'manual')""", (DATE, mid, real, 0, real, 0))
        n_raw += 1
    for name, real, can_qty, can_cnt in parse_sub():
        mid, new = find_or_create(name, "sub")
        if new:
            created.append(("sub", name))
        con.execute("""INSERT OR REPLACE INTO material_daily
            (date, material_id, prev_qty, in_qty, real_qty, used_qty, src)
            VALUES(?,?,?,?,?,?,'manual')""", (DATE, mid, real, 0, real, 0))
        n_sub += 1
        # 생산가능 환산: 수량 = 재고 × mult, 횟수 = 수량 ÷ per (시트 수식값에서 역산)
        if can_qty and real and can_qty > 0 and real > 0:
            mult = can_qty / real
            per = can_qty / can_cnt if can_cnt and can_cnt > 0 else None
            con.execute("UPDATE material SET prod_mult=?, prod_per=COALESCE(?, prod_per) WHERE id=?",
                        (mult, per, mid))
            n_mult += 1
    con.execute("INSERT OR IGNORE INTO day_record(date) VALUES(?)", (DATE,))
    print(f"임포트: 원재료 {n_raw}종 · 부재료 {n_sub}종 (신규 생성 {len(created)}종) · 환산계수 갱신 {n_mult}종")
    for k, n in created:
        print(f"  신규: [{'원' if k == 'raw' else '부'}] {n}")

    # 7/8 자동차감 행 재계산: 전일재고 = 새 7/7 실재고 (없으면 0), 실재고 = 전일 + 입고 − 사용
    n_fix = 0
    for r in con.execute("SELECT * FROM material_daily WHERE date>? AND src='auto' ORDER BY date",
                         ("2026-07-07",)).fetchall():
        prev_row = con.execute("""SELECT real_qty FROM material_daily
            WHERE material_id=? AND date<? ORDER BY date DESC LIMIT 1""",
                               (r["material_id"], r["date"])).fetchone()
        prev = prev_row["real_qty"] if prev_row else 0.0
        real = prev + r["in_qty"] - r["used_qty"]
        if abs(prev - r["prev_qty"]) > 1e-9 or abs(real - r["real_qty"]) > 1e-9:
            con.execute("UPDATE material_daily SET prev_qty=?, real_qty=? WHERE id=?",
                        (prev, real, r["id"]))
            n_fix += 1
    print(f"7/8 자동차감 재계산: {n_fix}건 갱신")

    con.execute("INSERT INTO audit_log(action,detail) VALUES(?,?)",
                ("import_stock_0707",
                 f"실사 재설정 raw {n_raw} sub {n_sub} 신규 {len(created)} 7/8보정 {n_fix}"))
    con.commit()

    # 검증 출력
    print("--- 검증 (샘플 5종) ---")
    for r in con.execute("""SELECT m.name, m.kind, md.real_qty FROM material_daily md
        JOIN material m ON m.id=md.material_id WHERE md.date=? LIMIT 5""", (DATE,)):
        print(f"  {r['name']} [{r['kind']}] = {r['real_qty']}")
    con.close()


if __name__ == "__main__":
    main()
