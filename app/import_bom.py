# -*- coding: utf-8 -*-
"""doc/배합비/*.xlsx → bom 테이블 (기존 배합비 전체 교체).

- 각 파일 = 제품 1종의 배합 (첫 시트 = 현행, '1 배합' 컬럼 B/C만 사용)
- 여러 블록(반죽+토핑 등)은 자재별 합산
- 수량은 '1배합당 g' → 개당 환산: qty × 제품중량(g) ÷ 배합총량(g)
  (제품중량은 제품명의 '(45 g)' 표기에서 추출 — 중량 다른 변형 제품은 각자 환산)
- 자재 매칭: 정확일치 → 별칭표 → 유일 포함 → 신규 생성(note='배합비 유래')
"""
import io
import re
import sys
import glob
import datetime as dt
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from db import connect
from import_excel import mkey, pkey

BOM_DIR = Path(__file__).resolve().parent.parent / "doc" / "배합비"

# 파일명 키 → (제품 검색 토큰들, 시트 힌트)  · 토큰: 제품 pkey에 모두 포함되어야 매칭
FILE_MAP = {
    "e마트_오븐에구운소보로도넛": (["오븐에구운소보로"], None),        # 이마트+트레이더스 (동일 배합 준용)
    "넛파인_촉촉한 버터 도넛": (["The촉촉", "버터"], "버터"),
    "넛파인_촉촉한 아몬드 도넛": (["The촉촉", "아몬드"], "아몬드"),
    "넛파인_촉촉한 오리지널 도넛": (["The촉촉", "오리지널"], "오리지널"),
    "달광_골드치즈도넛": (["달광도넛", "골드치즈"], None),
    "달광_딸기도넛": (["달광도넛", "딸기"], None),
    "달광_바나나도넛": (["달광도넛", "바나나"], None),
    "달광_아몬드도넛": (["달광도넛", "아몬드"], None),
    "달광_초코도넛": (["달광도넛", "초코"], None),
    "맘스_우리밀 모카소보로": (["맘스", "모카"], None),
    "맘스_우리밀 소보로도넛": (["맘스", "우리밀소보로도넛"], None),
    "맘스_우리밀 카카오도넛": (["맘스", "카카오"], None),
    "맘스_우리밀도넛": (["맘스우리밀도넛"], None),
    "소보로 15_바나나 소보로": (["소보로", "바나나"], None),
    "소보로 15_블루베리 소보로": (["소보로", "블루베리"], None),
    "소보로 15_치즈 소보로": (["소보로", "골드치즈"], None),
    "소보로 15_코코넛 소보로": (["소보로", "코코넛"], None),
    "소보로 15_코코아 소보로": (["소보로", "코코아"], None),
    "파_골드치즈 파운드": (["파운드", "치즈"], None),   # '치즈 파운드' + 빵아빵아 골드치즈
    "파_소보로 파운드": (["파운드", "소보로"], None),
    "파_아몬드 파운드": (["파운드", "아몬드"], None),
    "푸르나이_통우리밀 소보로도넛": (["푸르나이", "소보로"], None),
    "푸르나이_통우리밀 초코도넛": (["푸르나이", "초코"], None),
}
SKIP_FILES = {
    "260128_소보로도넛 배합비율(5광)": "소배합 참고자료(총 2.7kg) — 대상 제품 특정 불가",
    "블루베리 파운드": "해당 제품이 수불부에 없음(소배합 참고자료)",
}
# 배합비 표기 → DB 자재명 별칭
ALIAS = {
    "중력분": "밀가루 중력분", "코트롤": "코트롤티에이치", "엑시드C": "엑시드-C",
    "소비톨": "D-소비톨액", "솔비톨": "D-소비톨액",
    "1차전란액": "전란액", "2차전란액": "전란액", "계란": "전란액",
    "베이킹파우더": "베이킹파우더포뮬러2", "밀크버터후레바": "밀키버터후레바",
    "우리밀밀가루": "우리밀 밀가루", "통우리밀": "통밀가루",
    "커피향": "커피후레바(향)", "커피분말": "맥스웰화인(커피분말)", "커피레진": "커피씨씨(레진)",
    "바나나레진": "바나나씨씨", "블루베리레진": "블루베리 씨씨", "블루베리": "블루베리 씨씨",
    "딸기레진": "딸기씨씨", "혼합치즈": "치즈혼합분말 1", "혼합치즈분말": "치즈혼합분말 1",
    "치즈혼합분말": "치즈혼합분말 1", "코코아분말": "코코아파우다",
    "국산버터": "버터", "소금": "정제소금", "유기농설탕": "이티자황설탕",
}
WEIGHT_PAT = re.compile(r"\(([\d.]+)\s*g\)", re.I)
# 무게 표기 없는 정식 제품 (파운드 3종은 2026-07-07 통합 개명으로 무게 미표기)
NO_WEIGHT_OK = {"치즈 파운드", "아몬드 파운드", "소보로 파운드"}

# 반죽량.xlsx: 파일 → (도우배합 수율, 토핑배합 수율)  [개/1배합]
YIELDS = {
    "e마트_오븐에구운소보로도넛": (8753.1, 9821.2),
    "넛파인_촉촉한 버터 도넛": (6131.3, None),
    "넛파인_촉촉한 아몬드 도넛": (7063.6, None),
    "넛파인_촉촉한 오리지널 도넛": (6622.1, None),
    "달광_골드치즈도넛": (6051.2, None),
    "달광_딸기도넛": (6025.3, None),
    "달광_바나나도넛": (5926.3, None),
    "달광_아몬드도넛": (6056.3, None),
    "달광_초코도넛": (6253.0, None),
    "맘스_우리밀 모카소보로": (8668.2, 9677.8),
    "맘스_우리밀 소보로도넛": (8487.9, 9442.2),
    "맘스_우리밀 카카오도넛": (6157.7, None),
    "맘스_우리밀도넛": (5632.4, None),
    "소보로 15_바나나 소보로": (9149.4, 7884.4),
    "소보로 15_블루베리 소보로": (8477.9, 7966.9),
    "소보로 15_치즈 소보로": (8895.2, 8168.8),
    "소보로 15_코코넛 소보로": (8925.5, 9281.3),
    "소보로 15_코코아 소보로": (9053.3, 8359.4),
    "파_골드치즈 파운드": (4751.7, None),
    "파_소보로 파운드": (5735.7, 9442.2),
    "파_아몬드 파운드": (4891.5, None),
    "푸르나이_통우리밀 소보로도넛": (8812.5, 8666.7),
    "푸르나이_통우리밀 초코도넛": (6366.0, None),
}


def parse_recipe(path, sheet_hint):
    """블록(반죽/토핑)별로 분리 파싱 → [(block_items{name:qty}, block_total), ...]"""
    wb = load_workbook(path, read_only=True, data_only=True)
    sn = wb.sheetnames[0]
    if sheet_hint:
        for n in wb.sheetnames:
            if sheet_hint in n.replace(" ", ""):
                sn = n
                break
    blocks = []
    cur = None
    for row in wb[sn].iter_rows(min_row=1, max_col=3, values_only=True):
        b, c = row[1], row[2]
        if b is None or not str(b).strip():
            continue
        t = re.sub(r"\s+", " ", str(b).strip())
        if "배합" in t:                       # 블록 헤더 → 새 블록 시작
            cur = {}
            blocks.append(cur)
            continue
        if isinstance(c, (int, float)) and c > 0:
            if cur is None:
                cur = {}
                blocks.append(cur)
            cur[t] = cur.get(t, 0.0) + float(c)
    wb.close()
    return sn, [(blk, sum(blk.values())) for blk in blocks if blk]


def main():
    con = connect()
    products = [dict(r) for r in con.execute(
        "SELECT id, name FROM product ORDER BY sort, id")]
    raw_mats = {mkey(r["name"]): (r["id"], r["name"]) for r in con.execute(
        "SELECT id, name FROM material WHERE kind='raw'")}

    def find_material(name):
        k = mkey(name)
        if k in raw_mats:
            return raw_mats[k][0]
        if name in ALIAS or k in {mkey(a) for a in ALIAS}:
            target = ALIAS.get(name) or next(v for a, v in ALIAS.items() if mkey(a) == k)
            tk = mkey(target)
            if tk in raw_mats:
                return raw_mats[tk][0]
        cands = [v for kk, v in raw_mats.items() if k and k in kk]
        if len(cands) == 1:
            return cands[0][0]
        # 신규 생성
        con.execute("INSERT INTO material(kind,name,unit,note) VALUES('raw',?, 'kg', '배합비 유래')",
                    (name,))
        mid = con.execute("SELECT id FROM material WHERE kind='raw' AND name=?",
                          (name,)).fetchone()[0]
        raw_mats[mkey(name)] = (mid, name)
        created.append(name)
        return mid

    con.execute("DELETE FROM bom")
    created, report, skipped = [], [], []
    today = dt.date.today().isoformat()

    for path in sorted(glob.glob(str(BOM_DIR / "*.xlsx"))):
        stem = Path(path).stem.rstrip("-").strip()
        if stem in SKIP_FILES:
            skipped.append(f"{stem} — {SKIP_FILES[stem]}")
            continue
        entry = FILE_MAP.get(stem)
        if not entry:
            skipped.append(f"{stem} — 매핑 없음")
            continue
        tokens, hint = entry
        sn, blocks = parse_recipe(path, hint)
        if not blocks:
            skipped.append(f"{stem} — 파싱 실패")
            continue
        yields = YIELDS.get(stem)
        if not yields:
            skipped.append(f"{stem} — 반죽량(수율) 정보 없음")
            continue
        targets = [p for p in products
                   if all(pkey(t).lower() in pkey(p["name"]).lower() for t in tokens)]
        targets = [p for p in targets
                   if WEIGHT_PAT.search(p["name"]) or p["name"] in NO_WEIGHT_OK]
        if not targets:
            skipped.append(f"{stem} — 대상 제품 없음 (토큰 {tokens})")
            continue
        for p in targets:
            # 제품에 배합당 생산수량(도우 기준) 기록
            con.execute("UPDATE product SET batch_yield=? WHERE id=?", (yields[0], p["id"]))
            n = 0
            for bi, (items, total) in enumerate(blocks):
                y = yields[bi] if bi < len(yields) and yields[bi] else yields[0]
                blabel = "반죽" if bi == 0 else "토핑"
                # 토핑 배합은 별도 제품('<이름> 토핑')으로 등록 (2026-07-08 사용자 확정 구조)
                target_pid = p["id"]
                if blabel == "토핑":
                    base = re.sub(r"\s*\([^)]*g\)\s*", " ", p["name"]).strip()
                    tname = re.sub(r"\s+", " ", base + " 토핑")
                    row = con.execute("SELECT id FROM product WHERE name=?", (tname,)).fetchone()
                    if row:
                        target_pid = row["id"] if hasattr(row, "keys") else row[0]
                    else:
                        con.execute("""INSERT INTO product(name, category, note, status)
                            VALUES(?, '', '토핑 반제품 (배합 분리)', '판매중')""", (tname,))
                        target_pid = con.execute("SELECT id FROM product WHERE name=?",
                                                 (tname,)).fetchone()[0]
                    con.execute("UPDATE product SET batch_yield=? WHERE id=?", (y, target_pid))
                for mat_name, qty in items.items():
                    mid = find_material(mat_name)
                    # 블록별 개별 행: batch_qty = 엑셀 원본 수치 그대로 (표시용),
                    # qty_per_unit = 블록배합량 ÷ 블록수율 (일일 입력·이론사용 계산용).
                    # 같은 블록 안의 별칭 합류(1차+2차 전란액)만 합산, 반죽/토핑 교차 합산은 안 함.
                    cur = con.execute(
                        """SELECT id, qty_per_unit, batch_qty FROM bom
                           WHERE product_id=? AND material_id=? AND block=?""",
                        (target_pid, mid, blabel)).fetchone()
                    if cur:
                        con.execute("""UPDATE bom SET qty_per_unit=?, batch_qty=?,
                            note='별칭 합산 (1차+2차 등)' WHERE id=?""",
                                    (cur["qty_per_unit"] + qty / y,
                                     cur["batch_qty"] + qty, cur["id"]))
                    else:
                        con.execute("""INSERT INTO bom
                            (product_id, material_id, qty_per_unit, unit,
                             block, batch_qty, block_yield, effective_from, note)
                            VALUES(?,?,?,?,?,?,?,?,?)""",
                                    (target_pid, mid, qty / y, "g",
                                     blabel, qty, y, today, ""))
                    n += 1
            report.append(f"  {stem} → {p['name']} ({n}자재 · 배합수율 {yields[0]:g}개"
                          f"{' + 토핑 ' + format(yields[1], 'g') + '개' if len(yields) > 1 and yields[1] else ''})")
    con.execute("INSERT INTO audit_log(action,detail) VALUES('import_bom','배합비 엑셀 전체 교체')")
    con.commit()

    print("=== 배합비 등록 ===")
    for r in report:
        print(r)
    nb = con.execute("SELECT COUNT(*) FROM bom").fetchone()[0]
    np = con.execute("SELECT COUNT(DISTINCT product_id) FROM bom").fetchone()[0]
    print(f"\n합계: 제품 {np}종, 배합 행 {nb}개")
    if created:
        print("신규 생성 자재('배합비 유래'):", " · ".join(created))
    if skipped:
        print("건너뜀:")
        for s in skipped:
            print("  -", s)
    con.close()


if __name__ == "__main__":
    main()
